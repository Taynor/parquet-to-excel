import json
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Border, Side, Alignment
import pandas as pd
import excel_config
from copy import copy

class ExcelCustomStyler:
    def __init__(self, json_specification, json_styling, excel_file, default_sheet_name, 
    parquet_load_file, default_load_sheet, worksheets, content_bold_title_large_font, 
    content_title_large_font, content_bold_title_small_font, content_bold_title_large_alignment_style,
    content_value_medium_font, child_title_small_font, child_bold_title_small_font,
    child_values_small_font, hyperlink_underline_style, thin_style, thin_borders_side_style, 
    thin_borders_top_style, thin_borders_bottom_style, thin_borders_full_style,
    thin_borders_top_bottom_style, thin_borders_top_bottom_right_style, thin_borders_top_bottom_left_style):
        
        #these fields are available for the user to manipulate
        self.json_specification = json_specification
        self.json_styling = json_styling
        self.excel_file = excel_file
        self.default_sheet_name = default_sheet_name
        self.parquet_load_file = parquet_load_file

        #these fields are hidden and are for PRIVATE use only
        self.worksheets = worksheets
        self.default_load_sheet = default_load_sheet
        self.content_bold_title_large_font = content_bold_title_large_font
        self.content_title_large_font = content_title_large_font
        self.content_bold_title_small_font = content_bold_title_small_font
        self.content_bold_title_large_alignment_style = content_bold_title_large_alignment_style
        self.content_value_medium_font = content_value_medium_font
        self.child_title_small_font = child_title_small_font
        self.child_bold_title_small_font = child_bold_title_small_font
        self.child_values_small_font = child_values_small_font
        self.hyperlink_underline_style = hyperlink_underline_style
        self.thin_style = thin_style
        self.thin_borders_side_style = thin_borders_side_style
        self.thin_borders_top_style = thin_borders_top_style
        self.thin_borders_bottom_style = thin_borders_bottom_style
        self.thin_borders_full_style = thin_borders_full_style
        self.thin_borders_top_bottom_style = thin_borders_top_bottom_style
        self.thin_borders_top_bottom_right_style = thin_borders_top_bottom_right_style
        self.thin_borders_top_bottom_left_style = thin_borders_top_bottom_left_style

    #property get and set for json specification path value  
    @property
    def json_specification(self):
        return self.__json_specification
    @json_specification.setter
    def json_specification(self, value):
        self.__json_specification = value
    
    #property get and set for json styling path value
    @property
    def json_styling(self):
        return self.__json_styling
    @json_styling.setter
    def json_styling(self, value):
        self.__json_sytyling = value

    #property get and set for excel file value  
    @property
    def excel_file(self):
        return self.__excel_file
    @excel_file.setter
    def excel_file(self, value):
        self.__excel_file = value    

    #property get and set for the default worksheet value
    @property
    def default_sheet_name(self):
        return self.__default_sheet_name
    @default_sheet_name.setter
    def default_sheet_name(self, value):
        self.__default_sheet_name = value 

    #property get and set for parquet file value 
    @property
    def parquet_load_file(self):
        return self.__parquet_file
    @parquet_load_file.setter
    def parquet_file(self, value):
        self.__parquet_file = value 

    #property get for parquet default load sheet value
    @property
    def default_load_sheet(self):
        return self.__default_load_sheet
    @default_load_sheet.setter
    def default_load_sheet(self, value):
        self.__default_load_sheet = value

    #property get and set for worksheets value
    @property 
    def worksheets(self):
        return self.__worksheets
    @worksheets.setter
    def worksheets(self, value):
        self.__worksheets = value

    #property get and set for content bold title large font value
    @property 
    def content_bold_title_large_font(self):
        return self.__content_bold_title_large_font
    @content_bold_title_large_font.setter
    def content_bold_title_large_font(self, value):
        self.__content_bold_title_large_font = value

    #property get and set for content title large font value
    @property 
    def content_title_large_font(self):
        return self.__content_title_large_font
    @content_title_large_font.setter
    def content_title_large_font(self, value):
        self.__content_title_large_font = value

    #property get and set for content bold title small font value
    @property 
    def content_bold_title_small_font(self):
        return self.__content_bold_title_small_font
    @content_bold_title_small_font.setter
    def content_bold_title_small_font(self, value):
        self.__content_bold_title_small_font = value 

    #property get and set for content bold title large alignment style value
    @property 
    def content_bold_title_large_alignment_style(self):
        return self.__content_bold_title_large_alignment_style
    @content_bold_title_large_alignment_style.setter
    def content_bold_title_large_alignment_style(self, value):
        self.__content_bold_title_large_alignment_style = value   

    #property get and set for content value medium font value
    @property 
    def content_value_medium_font(self):
        return self.__content_value_medium_font
    @content_value_medium_font.setter
    def content_value_medium_font(self, value):
        self.__content_value_medium_font = value  

    #property get and set for child title small font value
    @property 
    def child_title_small_font(self):
        return self.__child_title_small_font
    @child_title_small_font.setter
    def child_title_small_font(self, value):
        self.__child_title_small_font = value    

    #property get and set for child bold title small font value
    @property 
    def child_bold_title_small_font(self):
        return self.__child_bold_title_small_font
    @child_bold_title_small_font.setter
    def child_bold_title_small_font(self, value):
        self.__child_bold_title_small_font = value 

    #property get and set for child values small font value
    @property 
    def child_values_small_font(self):
        return self.__child_values_small_font
    @child_values_small_font.setter
    def child_values_small_font(self, value):
        self.__child_values_small_font = value

    #property get and set for hyperlink underline style value
    @property 
    def hyperlink_underline_style(self):
        return self.__hyperlink_underline_style
    @hyperlink_underline_style.setter
    def hyperlink_underline_style(self, value):
        self.__hyperlink_underline_style = value

    #property get and set for thin style value
    @property 
    def thin_style(self):
        return self.__thin_style
    @thin_style.setter
    def thin_style(self, value):
        self.__thin_style = value   

    #property get and set for thin borders side style value
    @property 
    def thin_borders_side_style(self):
        return self.__thin_borders_side_style
    @thin_borders_side_style.setter
    def thin_borders_side_style(self, value):
        self.__thin_borders_side_style = value      

    #property get and set for thin borders top style value
    @property 
    def thin_borders_top_style(self):
        return self.__thin_borders_top_style
    @thin_borders_top_style.setter
    def thin_borders_top_style(self, value):
        self.__thin_borders_top_style = value       

    #property get and set for thin borders bottom style value
    @property 
    def thin_borders_bottom_style(self):
        return self.__thin_borders_bottom_style
    @thin_borders_bottom_style.setter
    def thin_borders_bottom_style(self, value):
        self.__thin_borders_bottom_style = value 

    #property get and set for thin borders full style value
    @property 
    def thin_borders_full_style(self):
        return self.__thin_borders_full_style
    @thin_borders_full_style.setter
    def thin_borders_full_style(self, value):
        self.__thin_borders_full_style  

    #property get and set for thin borders top bottom style value
    @property 
    def thin_borders_top_bottom_style(self):
        return self.__thin_borders_top_bottom_style
    @thin_borders_top_bottom_style.setter
    def thin_borders_top_bottom_style(self, value):
        self.__thin_borders_top_bottom_style   

    #property get and set for thin borders top bottom right style value
    @property 
    def thin_borders_top_bottom_right_style(self):
        return self.__thin_borders_top_bottom_right_style
    @thin_borders_top_bottom_right_style.setter
    def thin_borders_top_bottom_right_style(self, value):
        self.__thin_borders_top_bottom_right_style        

    #property get and set for thin borders top bottom left style value
    @property 
    def thin_borders_top_bottom_left_style(self):
        return self.__thin_borders_top_bottom_left_style
    @thin_borders_top_bottom_left_style.setter
    def thin_borders_top_bottom_left_style(self, value):
        self.__thin_borders_top_bottom_left_style                                                       

    #read in the style sheet to load in the global variables for styling
    def read_style_sheet(json_styling):
        
        #pass the parameters to class variables for reuse
        ExcelCustomStyler.json_styling = json_styling

        #load styles from json_styles
        with open(json_styling) as jst:
            style_config = json.load(jst)

            #load the styles for large content bold title labels
            for content_bold_title_large in style_config["custom_content_bold_title_large"]:
                content_bold_title_large_font = Font(name=content_bold_title_large['font_name'],
                size=content_bold_title_large['font_size'],
                bold=content_bold_title_large['bold'],
                color=content_bold_title_large['font_colour'])

            #load the styles for large content title values
            for content_title_large in style_config["custom_content_title_large"]:   
                content_title_large_font = Font(name=content_title_large['font_name'],
                size=content_title_large['font_size'],
                color=content_title_large['font_colour'])

            #load the style for small content bold title values
            for content_bold_title_small in style_config["custom_content_bold_title_small"]:
                content_bold_title_small_font = Font(name=content_bold_title_small['font_name'],
                size=content_bold_title_small['font_size'],
                bold=content_bold_title_large['bold'],
                color=content_bold_title_small['font_colour'])     

            #load the styles for large content bold title labels alignment
            for content_bold_title_large_alignment in style_config["custom_alignment_content_bold_title_large"]:
                content_bold_title_large_alignment_style = Alignment(horizontal=content_bold_title_large_alignment['horizontal'],
                vertical=content_bold_title_large_alignment['vertical'])   

            #load the styles for medium content values
            for content_value_medium in style_config["custom_content_sheet_font"]:
                content_value_medium_font = Font(name=content_value_medium['font_name'],
                size=content_value_medium['font_size'],
                color=content_value_medium['font_colour'])   

            #load the styles for the small child title values
            for child_title_small in style_config["custom_child_title_small"]:
                child_title_small_font = Font(name=child_title_small['font_name'],
                size=child_title_small['font_size'],
                color=child_title_small['font_colour'])    

            #load the styles for the small child bold title values
            for child_bold_title_small in style_config["custom_child_bold_title_small"]:
                child_bold_title_small_font = Font(name=child_bold_title_small['font_name'],
                size=child_bold_title_small['font_size'],
                bold=child_bold_title_small['bold'],
                color=child_bold_title_small['font_colour'])   

            #load the styles for the small child values
            for child_values_small in style_config["custom_child_values_small"]:
                child_values_small_font = Font(name=child_values_small['font_name'],
                size=child_values_small['font_size'],
                color=child_values_small['font_colour'])

            #load the styles for the small bold child values
            for child_bold_values_small in style_config["custom_child_bold_values_small"]:
                child_bold_values_small_font = Font(name=child_bold_values_small['font_name'],
                size=child_bold_values_small['font_size'],
                bold=child_bold_values_small['bold'],
                color=child_bold_values_small['font_colour'])    

            #load the styles for the child sheet to content sheet hyperlink
            for hyperlink_underline in style_config["custom_hyperlink_underline"]:
                hyperlink_underline_style = Font(name=hyperlink_underline['font_name'],
                size=hyperlink_underline['font_size'],
                underline=hyperlink_underline['underline'],
                color=hyperlink_underline['font_colour'])
            
            #load the styles for the base sheet borders
            for thin_borders in style_config["custom_thin_borders"]:             
                thin_style = Side(border_style=thin_borders['border_width'],
                color=thin_borders['border_colour'])
                thin_borders_side_style = Border(right=thin_style, left=thin_style)
                thin_borders_top_style = Border(top=thin_style, right=thin_style, left=thin_style)
                thin_borders_bottom_style = Border(bottom=thin_style, right=thin_style, left=thin_style)  
                thin_borders_full_style = Border(top=thin_style, bottom=thin_style, right=thin_style, left=thin_style)
                thin_borders_top_bottom_style = Border(top=thin_style, bottom=thin_style)
                thin_borders_top_bottom_right_style = Border(top=thin_style, bottom=thin_style, right=thin_style)
                thin_borders_top_bottom_left_style = Border(top=thin_style, bottom=thin_style, left=thin_style)
            
            #pass the parameters to class variables for reuse
            ExcelCustomStyler.content_bold_title_large_font = content_bold_title_large_font
            ExcelCustomStyler.content_title_large_font = content_title_large_font
            ExcelCustomStyler.content_bold_title_small_font = content_bold_title_small_font
            ExcelCustomStyler.content_bold_title_large_alignment_style = content_bold_title_large_alignment_style
            ExcelCustomStyler.content_value_medium_font = content_value_medium_font
            ExcelCustomStyler.child_title_small_font = child_title_small_font
            ExcelCustomStyler.child_bold_title_small_font = child_bold_title_small_font
            ExcelCustomStyler.child_values_small_font = child_values_small_font
            ExcelCustomStyler.child_bold_values_small_font = child_bold_values_small_font
            ExcelCustomStyler.hyperlink_underline_style = hyperlink_underline_style
            ExcelCustomStyler.thin_style = thin_style
            ExcelCustomStyler.thin_borders_side_style = thin_borders_side_style
            ExcelCustomStyler.thin_borders_top_style = thin_borders_top_style
            ExcelCustomStyler.thin_borders_bottom_style = thin_borders_bottom_style
            ExcelCustomStyler.thin_borders_full_style = thin_borders_full_style
            ExcelCustomStyler.thin_borders_top_bottom_style = thin_borders_top_bottom_style
            ExcelCustomStyler.thin_borders_top_bottom_right_style = thin_borders_top_bottom_right_style
            ExcelCustomStyler.thin_borders_top_bottom_left_style = thin_borders_top_bottom_left_style
    
    #takes the arguments to build the worksheets list in order for the styler
    #to style the worksheets in the dynamic list of worksheets
    def style_worksheets(excel_file, parquet_load_file, json_specification, json_styling, default_sheet_name = 'Sheet1'):

        #import the styling sheet
        ExcelCustomStyler.read_style_sheet(json_styling)
        
        #pass the parameters to class variables for reuse
        ExcelCustomStyler.excel_file = excel_file
        ExcelCustomStyler.parquet_load_file = parquet_load_file
        ExcelCustomStyler.json_specification = json_specification
        ExcelCustomStyler.json_styling = json_styling

        #read the contents of the parquet loader that has the filter to where the data
        #should be loaded into the respective worksheets
        ExcelCustomStyler.read_parquet_loader(ExcelCustomStyler.parquet_load_file)

    #read parquet file that acts as a pointer to where the data needs to be loaded into
    def read_parquet_loader(parquet_load_file):
        
        #read the content of the parquet loading file
        ExcelCustomStyler.default_load_sheet = pd.read_parquet(parquet_load_file, engine='fastparquet')

        #set the filter column read in from the parquet
        ExcelCustomStyler.create_worksheet_list()
    
    #creates a list of worksheets dynamically created from the parquet filter column
    def create_worksheet_list():
        
        #create an empty to append the column values from the parquet filter
        worksheets = []
         
        #apply the filter to the parquet data loader file to build the list of
        #respective worksheets that will have it's data loaded into
        #there is a bug that means the filter column cannot be passed in as variable
        #it needs to be hardcoded in as an attribute of the dataframe
        ExcelCustomStyler.filter_column = ExcelCustomStyler.default_load_sheet.Question
        for fc in ExcelCustomStyler.filter_column:
            worksheets.append(fc)

        #pass the parameters to class variables for reuse   
        ExcelCustomStyler.worksheets = worksheets

        #load the excel file with the data loaded to start applying styling
        #ExcelCustomStyler.load_excel_worksheet(ExcelCustomStyler.excel_file, ExcelCustomStyler.json_specification, ExcelCustomStyler.worksheets)
        ExcelCustomStyler.apply_content_worksheet_style(ExcelCustomStyler.worksheets, ExcelCustomStyler.excel_file, ExcelCustomStyler.default_sheet_name, ExcelCustomStyler.json_specification)

    #apply the styling for the content worksheet
    def apply_content_worksheet_style(worksheets, excel_file, default_sheet_name, json_specification):

        #load the workbook
        excel_workbook = load_workbook(excel_file)
        worksheet = excel_workbook.active

        #open up the spec sheet to grab custom content to add to the Content worksheet
        with open(json_specification) as jsp:
            spec_config = json.load(jsp)
            
            #Add the labels and titles
            for content_project_details in spec_config["content_project_details"]:
                worksheet[excel_config.content_sheet_labels[0]] = content_project_details['Project_Label']
                worksheet[excel_config.content_sheet_labels[1]] = content_project_details['Project_Name']
                worksheet[excel_config.content_sheet_labels[2]] = content_project_details['Wave_Label']
                worksheet[excel_config.content_sheet_labels[3]] = content_project_details['Wave_Name']
                worksheet[excel_config.content_sheet_labels[4]] = content_project_details['Fieldwork_Label']
                worksheet[excel_config.content_sheet_labels[5]] = content_project_details['Fieldwork_Name'] 

                #Add styling import
                #Add styling to the large formatting for the content sheet title labels
                worksheet[excel_config.content_sheet_labels[0]].font = ExcelCustomStyler.content_bold_title_large_font
                worksheet[excel_config.content_sheet_labels[2]].font = ExcelCustomStyler.content_bold_title_large_font
                worksheet[excel_config.content_sheet_labels[4]].font = ExcelCustomStyler.content_bold_title_large_font

                #Add styling to the large formatting for the content sheet title values
                worksheet[excel_config.content_sheet_labels[1]].font = ExcelCustomStyler.content_title_large_font
                worksheet[excel_config.content_sheet_labels[3]].font = ExcelCustomStyler.content_title_large_font
                worksheet[excel_config.content_sheet_labels[5]].font = ExcelCustomStyler.content_title_large_font

                #Add styling to the small formatting for the content sheet title values
                worksheet[excel_config.content_sheet_labels[6]].font = ExcelCustomStyler.content_bold_title_small_font
                worksheet[excel_config.content_sheet_labels[7]].font = ExcelCustomStyler.content_bold_title_small_font
                worksheet[excel_config.content_sheet_labels[8]].font = ExcelCustomStyler.content_bold_title_small_font  

                #Add freeze panes to the content sheet
                worksheet.freeze_panes = excel_config.content_sheet_labels[9]                           
        
        #Add content for child sheets
        excel_workbook.save(ExcelCustomStyler.excel_file)

        #execute the function to apply the French questions for each child work sheet
        ExcelCustomStyler.apply_french_questions_style(ExcelCustomStyler.worksheets, ExcelCustomStyler.excel_file, ExcelCustomStyler.default_sheet_name)

    #apply the content and styling for the French questions on each child worksheet
    def apply_french_questions_style(worksheets, excel_file, default_sheet_name):    

        #load the workbook
        excel_workbook = load_workbook(excel_file)
        worksheet = excel_workbook.active
        
        #create two lists designed to be used as the comparison statement
        #for loading the worksheet questions in for their respective worksheets
        cells_list = []
        cells_questions_list = []
        for row in worksheet.iter_rows(min_col=2, min_row=6, max_col=2):
            for cell in row:
                if cell.value != '' or cell.value != None:
                    current_cell_string = cell.value
                    new_cell_string = current_cell_string.split(".", 1)
                    substring_cell_string = new_cell_string[0]   
                    cells_list.append(substring_cell_string)
                    cells_questions_list.append(current_cell_string)
        
        #loop and grab values to copy into the child sheets for the French questions
        for row in worksheet.iter_rows(min_col=2, min_row=6, max_col=2):
            for cell in row:
                if cell.value != '' or cell.value != None:
                    for cl in cells_list:
                        for ws in worksheets:
                            if cl == ws and ws != 'Content':
                                worksheet = excel_workbook[ws[0:]]
                                worksheet.cell(row=3, column=2).value = cell.value
                                worksheet.merge_cells(excel_config.child_sheet_question_content[0])

        #Add content for child sheets
        excel_workbook.save(ExcelCustomStyler.excel_file)   

        #execute the function to apply the English questions for each child work sheet
        ExcelCustomStyler.apply_english_questions_style(ExcelCustomStyler.worksheets, ExcelCustomStyler.excel_file, ExcelCustomStyler.default_sheet_name)

    #apply the content and styling for the English questions on each child worksheet
    def apply_english_questions_style(worksheets, excel_file, default_sheet_name):     

        #load the workbook
        excel_workbook = load_workbook(excel_file)
        worksheet = excel_workbook.active

        #create two lists designed to be used as the comparison statement
        #for loading the worksheet questions in for their respective worksheets
        cells_list = []
        cells_questions_list = []
        for row in worksheet.iter_rows(min_col=2, min_row=6, max_col=2):
            for cell in row:
                if cell.value != '' or cell.value != None:
                    current_cell_string = cell.value
                    new_cell_string = current_cell_string.split(".", 1)
                    substring_cell_string = new_cell_string[0]   
                    cells_list.append(substring_cell_string)
                    cells_questions_list.append(current_cell_string)                     
        
        #loop and grab values to copy into the child sheets for the English questions
        for row in worksheet.iter_rows(min_col=3, min_row=6, max_col=3):
            for cell in row:
                if cell.value != '' or cell.value != None:
                    for cl in cells_list:
                        for ws in worksheets:
                            if cl == ws and ws != 'Content':
                                worksheet = excel_workbook[ws[0:]]
                                worksheet.cell(row=3, column=8).value = cell.value
                                worksheet.merge_cells(excel_config.child_sheet_question_content[1])
                   
        #Add content for child sheets
        excel_workbook.save(ExcelCustomStyler.excel_file) 

        #execute the applying the hyperlinks to the content worksheets with styling
        ExcelCustomStyler.apply_content_worksheet_hyperlinks(ExcelCustomStyler.worksheets, ExcelCustomStyler.excel_file, ExcelCustomStyler.default_sheet_name) 

    #apply the hyperlinks and styling to the content worksheet, to navigate to the 
    #child worksheets     
    def apply_content_worksheet_hyperlinks(worksheets, excel_file, default_sheet_name):     

        #load the workbook
        excel_workbook = load_workbook(excel_file)
        worksheet = excel_workbook.active                                                         

        #iterate through the content columns and apply the font styling to these rows
        #where there are no blank values or nulls
        for row in worksheet.iter_rows(min_col=1, min_row=6, max_col=1):

            #build the hyperlink locations for the child cells 
            for ws in worksheets:
                hyperlink_location = "#"+ws+"!A1"
                for cell in row:

                    #apply the hyperlink and styling if the cell value and hyperlink
                    #location matches
                    if cell.value != '' or cell.value != None:
                        cell.font = ExcelCustomStyler.hyperlink_underline_style
                        if cell.value == ws:                       
                            cell.hyperlink = hyperlink_location

        #add the hyperlink content
        excel_workbook.save(ExcelCustomStyler.excel_file)                       

        #execute the content worksheet default font styling
        ExcelCustomStyler.apply_content_worksheet_default_font(ExcelCustomStyler.worksheets, ExcelCustomStyler.excel_file, ExcelCustomStyler.default_sheet_name)                    

    #apply the default styling to the content worksheet font
    def apply_content_worksheet_default_font(worksheets, excel_file, default_sheet_name): 

        #load the workbook
        excel_workbook = load_workbook(excel_file)
        worksheet = excel_workbook.active                          
        
        #format the rest of the values in the content worksheet
        for row in worksheet.iter_rows(min_col=2, min_row=6, max_col=3):
            for cell in row:
                if cell.value != '' or cell.value != None:
                    cell.font = ExcelCustomStyler.content_value_medium_font 

        #Add content for child sheets
        excel_workbook.save(ExcelCustomStyler.excel_file)  

        #execute applying the child worksheet default project details
        ExcelCustomStyler.apply_child_worksheet_project_details(ExcelCustomStyler.worksheets, ExcelCustomStyler.excel_file, ExcelCustomStyler.default_sheet_name, ExcelCustomStyler.json_specification) 

    #apply the default child worksheet project details
    def apply_child_worksheet_project_details(worksheets, excel_file, default_sheet_name, json_specification):   

        #load the workbook
        excel_workbook = load_workbook(excel_file)
        worksheet = excel_workbook.active

        #open up the spec sheet to grab custom content to add to the Content worksheet
        with open(json_specification) as jsp:
            spec_config = json.load(jsp)                       

            #apply styling for the other worksheets
            for ws in worksheets:

                #apply the default styling to the other child worksheets including the base sheet   
                if ws != 'Content':
                    worksheet = excel_workbook[ws[0:]]
                    for child_project_details in spec_config['child_project_details']:

                        #add child sheet content
                        worksheet[excel_config.child_sheet_labels[0]] = child_project_details['Project_Name']
                        worksheet[excel_config.child_sheet_labels[1]] = child_project_details['Wave_Name']
                        worksheet[excel_config.child_sheet_labels[2]] = child_project_details['Fieldwork_Name']
                        worksheet[excel_config.child_sheet_labels[3]] = child_project_details['Total']
                        worksheet[excel_config.child_sheet_labels[19]] = child_project_details['Hyperlink_Value']

                        #set up the hyperlink for back to content sheet
                        worksheet[excel_config.child_sheet_labels[19]].font = ExcelCustomStyler.hyperlink_underline_style
                        worksheet[excel_config.child_sheet_labels[19]].hyperlink = child_project_details['Content_Sheet_Hyperlink'] 

                        #format the layout
                        worksheet.merge_cells(excel_config.child_sheet_labels[4])    
                        worksheet.freeze_panes = excel_config.child_sheet_labels[5]   

                        #add styling to child sheet labels
                        worksheet[excel_config.child_sheet_labels[0]].font = ExcelCustomStyler.child_title_small_font
                        worksheet[excel_config.child_sheet_labels[1]].font = ExcelCustomStyler.child_title_small_font
                        worksheet[excel_config.child_sheet_labels[2]].font = ExcelCustomStyler.child_title_small_font
                        worksheet[excel_config.child_sheet_labels[3]].font = ExcelCustomStyler.child_bold_title_small_font

                        #add styling to the child sheet question labels
                        worksheet[excel_config.child_sheet_question_labels[0]].font = ExcelCustomStyler.child_title_small_font
                        worksheet[excel_config.child_sheet_question_labels[1]].font = ExcelCustomStyler.child_title_small_font

                        #Add borders to the Total cell in child sheets
                        worksheet[excel_config.child_sheet_regions[13]].border = ExcelCustomStyler.thin_borders_bottom_style

                        #Add border to the child and base region columns
                        worksheet[excel_config.child_sheet_regions[0]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[1]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[2]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[3]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[4]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[5]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[6]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[7]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[8]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[9]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[10]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[11]].border = ExcelCustomStyler.thin_borders_full_style
                        worksheet[excel_config.child_sheet_regions[12]].border = ExcelCustomStyler.thin_borders_full_style

                        #Add top border above child sheet regions columns
                        worksheet[excel_config.child_sheet_regions[14]].border = ExcelCustomStyler.thin_borders_top_bottom_left_style
                        worksheet[excel_config.child_sheet_regions[15]].border = ExcelCustomStyler.thin_borders_top_bottom_style
                        worksheet[excel_config.child_sheet_regions[16]].border = ExcelCustomStyler.thin_borders_top_bottom_style
                        worksheet[excel_config.child_sheet_regions[17]].border = ExcelCustomStyler.thin_borders_top_bottom_style
                        worksheet[excel_config.child_sheet_regions[18]].border = ExcelCustomStyler.thin_borders_top_bottom_style
                        worksheet[excel_config.child_sheet_regions[19]].border = ExcelCustomStyler.thin_borders_top_bottom_style
                        worksheet[excel_config.child_sheet_regions[20]].border = ExcelCustomStyler.thin_borders_top_bottom_style
                        worksheet[excel_config.child_sheet_regions[21]].border = ExcelCustomStyler.thin_borders_top_bottom_style
                        worksheet[excel_config.child_sheet_regions[22]].border = ExcelCustomStyler.thin_borders_top_bottom_style
                        worksheet[excel_config.child_sheet_regions[23]].border = ExcelCustomStyler.thin_borders_top_bottom_style
                        worksheet[excel_config.child_sheet_regions[24]].border = ExcelCustomStyler.thin_borders_top_bottom_style
                        worksheet[excel_config.child_sheet_regions[25]].border = ExcelCustomStyler.thin_borders_top_bottom_style
                        worksheet[excel_config.child_sheet_regions[26]].border = ExcelCustomStyler.thin_borders_top_bottom_right_style  

                    #add styling to the totals column values
                    for row in worksheet.iter_rows(min_col=3, min_row=10, max_col=15, max_row=10):
                        for cell in row:
                            if cell.value != '' or cell.value != None:
                                cell.font = ExcelCustomStyler.child_bold_values_small_font
                    
                    #add styling to the region column titles
                    for row in worksheet.iter_rows(min_col=3, min_row=9, max_col=15, max_row=9):
                        for cell in row:
                            if cell.value != '' or cell.value != None:
                                cell.font = ExcelCustomStyler.child_values_small_font              

                #apply the styling to the child worksheets 
                if ws != 'B' and ws != 'Content':

                    #perform child_sheet styling
                    #this styling is for the content for the child sheets values
                    for row in worksheet.iter_rows(min_col=3, min_row=11, max_col=15):
                        for cell in row:
                            if cell.value != '' or cell.value != None:
                                cell.font = ExcelCustomStyler.child_values_small_font       

                #apply the default content to the base worksheet
                if ws == 'B':               
                    worksheet = excel_workbook[ws[0:]]
                    for base_respondents_labels in spec_config['base_respondents_labels']:
                        worksheet[excel_config.base_sheet_labels[0]] = base_respondents_labels['Country_AL']
                        worksheet[excel_config.base_sheet_labels[1]] = base_respondents_labels['Country_AL']
                        worksheet[excel_config.base_sheet_labels[2]] = base_respondents_labels['Country_AT']
                        worksheet[excel_config.base_sheet_labels[3]] = base_respondents_labels['Country_AT']
                        worksheet[excel_config.base_sheet_labels[4]] = base_respondents_labels['Country_BE']
                        worksheet[excel_config.base_sheet_labels[5]] = base_respondents_labels['Country_BE']
                        worksheet[excel_config.base_sheet_labels[6]] = base_respondents_labels['Country_BG']
                        worksheet[excel_config.base_sheet_labels[7]] = base_respondents_labels['Country_BG'] 
                        worksheet[excel_config.base_sheet_labels[8]] = base_respondents_labels['Country_BA']
                        worksheet[excel_config.base_sheet_labels[9]] = base_respondents_labels['Country_BA']
                        worksheet[excel_config.base_sheet_labels[10]] = base_respondents_labels['Country_CH']
                        worksheet[excel_config.base_sheet_labels[11]] = base_respondents_labels['Country_CH']  
                        worksheet[excel_config.base_sheet_labels[12]] = base_respondents_labels['Country_CY']
                        worksheet[excel_config.base_sheet_labels[13]] = base_respondents_labels['Country_CY']
                        worksheet[excel_config.base_sheet_labels[14]] = base_respondents_labels['Country_CY(TCC)']
                        worksheet[excel_config.base_sheet_labels[15]] = base_respondents_labels['Country_CY(TCC)']
                        worksheet[excel_config.base_sheet_labels[16]] = base_respondents_labels['Country_CZ']
                        worksheet[excel_config.base_sheet_labels[17]] = base_respondents_labels['Country_CZ']
                        worksheet[excel_config.base_sheet_labels[18]] = base_respondents_labels['Country_DE']
                        worksheet[excel_config.base_sheet_labels[19]] = base_respondents_labels['Country_DE']
                        worksheet[excel_config.base_sheet_labels[20]] = base_respondents_labels['Country_DK']
                        worksheet[excel_config.base_sheet_labels[21]] = base_respondents_labels['Country_DK']
                        worksheet[excel_config.base_sheet_labels[22]] = base_respondents_labels['Country_EE']
                        worksheet[excel_config.base_sheet_labels[23]] = base_respondents_labels['Country_EE']
                        worksheet[excel_config.base_sheet_labels[24]] = base_respondents_labels['Country_EL(GR)']
                        worksheet[excel_config.base_sheet_labels[25]] = base_respondents_labels['Country_EL(GR)']
                        worksheet[excel_config.base_sheet_labels[26]] = base_respondents_labels['Country_ES']
                        worksheet[excel_config.base_sheet_labels[27]] = base_respondents_labels['Country_ES']
                        worksheet[excel_config.base_sheet_labels[28]] = base_respondents_labels['Country_FI']
                        worksheet[excel_config.base_sheet_labels[29]] = base_respondents_labels['Country_FI']
                        worksheet[excel_config.base_sheet_labels[30]] = base_respondents_labels['Country_FR']
                        worksheet[excel_config.base_sheet_labels[31]] = base_respondents_labels['Country_FR']
                        worksheet[excel_config.base_sheet_labels[32]] = base_respondents_labels['Country_HR']
                        worksheet[excel_config.base_sheet_labels[33]] = base_respondents_labels['Country_HR']
                        worksheet[excel_config.base_sheet_labels[34]] = base_respondents_labels['Country_HU']
                        worksheet[excel_config.base_sheet_labels[35]] = base_respondents_labels['Country_HU']
                        worksheet[excel_config.base_sheet_labels[36]] = base_respondents_labels['Country_IE']
                        worksheet[excel_config.base_sheet_labels[37]] = base_respondents_labels['Country_IE']
                        worksheet[excel_config.base_sheet_labels[38]] = base_respondents_labels['Country_IS']
                        worksheet[excel_config.base_sheet_labels[39]] = base_respondents_labels['Country_IS']
                        worksheet[excel_config.base_sheet_labels[40]] = base_respondents_labels['Country_IT']
                        worksheet[excel_config.base_sheet_labels[41]] = base_respondents_labels['Country_IT']
                        worksheet[excel_config.base_sheet_labels[42]] = base_respondents_labels['Country_KV']
                        worksheet[excel_config.base_sheet_labels[43]] = base_respondents_labels['Country_KV']
                        worksheet[excel_config.base_sheet_labels[44]] = base_respondents_labels['Country_LT']
                        worksheet[excel_config.base_sheet_labels[45]] = base_respondents_labels['Country_LT']
                        worksheet[excel_config.base_sheet_labels[46]] = base_respondents_labels['Country_LU']
                        worksheet[excel_config.base_sheet_labels[47]] = base_respondents_labels['Country_LU']
                        worksheet[excel_config.base_sheet_labels[48]] = base_respondents_labels['Country_LV']
                        worksheet[excel_config.base_sheet_labels[49]] = base_respondents_labels['Country_LV']
                        worksheet[excel_config.base_sheet_labels[50]] = base_respondents_labels['Country_ME']
                        worksheet[excel_config.base_sheet_labels[51]] = base_respondents_labels['Country_ME']
                        worksheet[excel_config.base_sheet_labels[52]] = base_respondents_labels['Country_MK']
                        worksheet[excel_config.base_sheet_labels[53]] = base_respondents_labels['Country_MK']
                        worksheet[excel_config.base_sheet_labels[54]] = base_respondents_labels['Country_MT']
                        worksheet[excel_config.base_sheet_labels[55]] = base_respondents_labels['Country_MT']
                        worksheet[excel_config.base_sheet_labels[56]] = base_respondents_labels['Country_NL']
                        worksheet[excel_config.base_sheet_labels[57]] = base_respondents_labels['Country_NL']
                        worksheet[excel_config.base_sheet_labels[58]] = base_respondents_labels['Country_NO']
                        worksheet[excel_config.base_sheet_labels[59]] = base_respondents_labels['Country_NO']
                        worksheet[excel_config.base_sheet_labels[60]] = base_respondents_labels['Country_PL']
                        worksheet[excel_config.base_sheet_labels[61]] = base_respondents_labels['Country_PL']
                        worksheet[excel_config.base_sheet_labels[62]] = base_respondents_labels['Country_PT']
                        worksheet[excel_config.base_sheet_labels[63]] = base_respondents_labels['Country_PT']
                        worksheet[excel_config.base_sheet_labels[64]] = base_respondents_labels['Country_RO']
                        worksheet[excel_config.base_sheet_labels[65]] = base_respondents_labels['Country_RO']
                        worksheet[excel_config.base_sheet_labels[66]] = base_respondents_labels['Country_RS']
                        worksheet[excel_config.base_sheet_labels[67]] = base_respondents_labels['Country_RS']
                        worksheet[excel_config.base_sheet_labels[68]] = base_respondents_labels['Country_SE']
                        worksheet[excel_config.base_sheet_labels[69]] = base_respondents_labels['Country_SE']
                        worksheet[excel_config.base_sheet_labels[70]] = base_respondents_labels['Country_SI']
                        worksheet[excel_config.base_sheet_labels[71]] = base_respondents_labels['Country_SI']
                        worksheet[excel_config.base_sheet_labels[72]] = base_respondents_labels['Country_SK']
                        worksheet[excel_config.base_sheet_labels[73]] = base_respondents_labels['Country_SK']
                        worksheet[excel_config.base_sheet_labels[74]] = base_respondents_labels['Country_TR']
                        worksheet[excel_config.base_sheet_labels[75]] = base_respondents_labels['Country_TR']
                        worksheet[excel_config.base_sheet_labels[76]] = base_respondents_labels['Country_UK']
                        worksheet[excel_config.base_sheet_labels[77]] = base_respondents_labels['Country_UK']
                        worksheet[excel_config.base_sheet_labels[78]] = base_respondents_labels['Country_US']
                        worksheet[excel_config.base_sheet_labels[79]] = base_respondents_labels['Country_US']
                        worksheet[excel_config.base_sheet_labels[80]] = base_respondents_labels['Country_WG']
                        worksheet[excel_config.base_sheet_labels[81]] = base_respondents_labels['Country_WG']
                        worksheet[excel_config.base_sheet_labels[82]] = base_respondents_labels['Country_EG']
                        worksheet[excel_config.base_sheet_labels[83]] = base_respondents_labels['Country_EG']

                        #Add borders styling to region country rows
                        worksheet[excel_config.base_sheet_labels[0]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[1]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[2]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[3]].border = ExcelCustomStyler.thin_borders_side_style 
                        worksheet[excel_config.base_sheet_labels[4]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[5]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[6]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[7]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[8]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[9]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[10]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[11]].border = ExcelCustomStyler.thin_borders_side_style 
                        worksheet[excel_config.base_sheet_labels[12]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[13]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[14]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[15]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[16]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[17]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[18]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[19]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[20]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[21]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[22]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[23]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[24]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[25]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[26]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[27]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[28]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[29]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[30]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[31]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[32]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[33]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[34]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[35]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[36]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[37]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[38]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[39]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[40]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[41]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[42]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[43]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[44]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[45]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[46]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[47]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[48]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[49]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[50]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[51]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[52]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[53]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[54]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[55]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[56]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[57]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[58]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[59]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[60]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[61]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[62]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[63]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[64]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[65]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[66]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[67]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[68]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[69]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[70]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[71]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[72]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[73]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[74]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[75]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[76]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[77]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[78]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[79]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[80]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[81]].border = ExcelCustomStyler.thin_borders_side_style
                        worksheet[excel_config.base_sheet_labels[82]].border = ExcelCustomStyler.thin_borders_top_style
                        worksheet[excel_config.base_sheet_labels[83]].border = ExcelCustomStyler.thin_borders_bottom_style    
                    
                    #perform base_sheet styling
                    #iterate through the content columns and apply the font styling to these rows
                    #where there are no blank values or nulls
                    #this styling is for the list of countries
                    for row in worksheet.iter_rows(min_col=2, min_row=12, max_col=2):
                        for cell in row:
                            if cell.value != '' or cell.value != None:
                                cell.font = ExcelCustomStyler.child_title_small_font

                    #this styling is for the content for the base sheet values
                    for row in worksheet.iter_rows(min_col=3, min_row=11, max_col=15):
                        for cell in row:
                            if cell.value != '' or cell.value != None:
                                cell.font = ExcelCustomStyler.child_values_small_font                                  
                
        #Add content for child sheets
        excel_workbook.save(ExcelCustomStyler.excel_file)  
