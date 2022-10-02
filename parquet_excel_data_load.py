from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import os
import pathlib
import glob
import data_load_config

class ParquetExcelDataLoad:
    def __init__(self, parquet_path, parquet_load_file, excel_file, parquet_list, parquet_subdirectories,
    worksheets, parquet_file_pattern, parquet_folders, parquet_filter, default_load_sheet, filter_column,
    default_sheet_name, custom_data_load):
        
        #these fields are available for the user to manipulate
        self.parquet_path = parquet_path
        self.parquet_file = parquet_file
        self.excel_file = excel_file
        self.parquet_filter = parquet_filter
        self.default_sheet_name = default_sheet_name
        self.custom_data_load = custom_data_load

        #these fields are hidden and are for PRIVATE use only
        self.parquet_list = parquet_list
        self.parquet_subdirectories = parquet_subdirectories
        self.worksheets = worksheets
        self.parquet_file_pattern = parquet_file_pattern
        self.parquet_folders = parquet_folders
        self.default_load_sheet = default_load_sheet
        self.filter_column = filter_column
    
    #property get and set for parquet path value
    @property
    def parquet_path(self):
        return self.__parquet_path
    @parquet_path.setter
    def parent_path(self, value):
        self.__parquet_path = value

    #property get and set for parquet file value 
    @property
    def parquet_load_file(self):
        return self.__parquet_file
    @parquet_load_file.setter
    def parquet_file(self, value):
        self.__parquet_file = value

    #property get and set for excel file value  
    @property
    def excel_file(self):
        return self.__excel_file
    @excel_file.setter
    def excel_file(self, value):
        self.__excel_file = value

    #property get and set for parquet filter value
    @property
    def parquet_filter(self):
        return self.__parquet_filter
    @parquet_filter.setter 
    def parquet_filter(self, value):
        self.__parquet_filter = value       

    #property get and set for parquet list value
    @property
    def parquet_list(self):
        return self.__parquet_list
    @parquet_list.setter 
    def parquet_list(self, value):
        self.__parquet_list = value

    #property get and set for parquet subdirectories value
    @property
    def parquet_subdirectories(self):
        return self.__parquet_subdirectories
    @parquet_subdirectories.setter
    def parquet_subdirectories(self, value):
        self.__parquet_subdirectories = value

    #property get and set for worksheets value
    @property 
    def worksheets(self):
        return self.__worksheets
    @worksheets.setter
    def worksheets(self, value):
        self.__worksheets = value
    
    #property get and set for parquet file pattern value
    @property
    def parquet_file_pattern(self):
        return self.parquet_file_pattern
    @parquet_file_pattern.setter 
    def parquet_file_pattern(self, value):
        self.parquet_file_pattern = value  

    #property get and set for parquet folder value     
    @property
    def parquet_folder(self):
        return self.__parquet_folder
    @parquet_folder.setter 
    def parquet_folder(self, value):
        self.__parquet_folder = value

    #property get for parquet default load sheet value
    @property
    def default_load_sheet(self):
        return self.__default_load_sheet
    @default_load_sheet.setter
    def default_load_sheet(self, value):
        self.__default_load_sheet = value     

    #property get for filter column value
    @property
    def filter_column(self):
        return self.__filter_column
    @filter_column.setter
    def filter_column(self, value):
        self.__filter_column = value 

    #property get and set for the default worksheet value
    @property
    def default_sheet_name(self):
        return self.__default_sheet_name
    @default_sheet_name.setter
    def default_sheet_name(self, value):
        self.__default_sheet_name = value

    #property get and set for the data custom loader value
    @property
    def custom_data_load(self):
        return self.__custom_data_load
    @custom_data_load.setter
    def custom_data_load(self, value):
        self.__custom_data_load = value                         

    #the main function that takes the arguments that will perform the data load
    #by taking the pararmeter values and using them as arguments for the downstream functions
    def load_parquet_data(parquet_load_file, excel_file, parquet_path, parquet_filter, default_sheet_name = 'Sheet1', custom_data_load = '', parquet_file_pattern = '\\*.parquet'):
        
        #local working variables required to build the list of objects
        parquet_subdirectories=[] 
        parquet_folders=[]
        parquet_list=[]

        #pass the parameters to class variables for reuse
        ParquetExcelDataLoad.parquet_load_file = parquet_load_file
        ParquetExcelDataLoad.parquet_filter = parquet_filter
        ParquetExcelDataLoad.excel_file = excel_file
        ParquetExcelDataLoad.default_sheet_name = default_sheet_name
        ParquetExcelDataLoad.parquet_load_file = parquet_load_file
        ParquetExcelDataLoad.custom_data_load = custom_data_load

        #create the list of parquet subdirectories
        for files in os.listdir(parquet_path):
            parquet_subdirectories.append(files)     

        #create the parquet list
        for parquet_directory in parquet_subdirectories:

            #the parquet subfolder to a string compare on and add to the list
            parquet_folder = parquet_path + parquet_directory
            parquet_folders.append(parquet_folder)

            #search through the sub folder and find the parquets
            parquets = glob.glob(parquet_folder + parquet_file_pattern)

            #append the list of parquets to the parquet list
            for p in parquets:
                parquet_list.append(p)  

            #reuse the parquet list object to load the data in
            ParquetExcelDataLoad.parquet_list = parquet_list    

        #read the contents of the parquet loader that has the filter to where the data
        #should be loaded into the respective worksheets
        ParquetExcelDataLoad.read_parquet_loader(ParquetExcelDataLoad.parquet_load_file)                      

    #read parquet file that acts as a pointer to where the data needs to be loaded into
    def read_parquet_loader(parquet_load_file):
        
        #read the content of the parquet loading file
        ParquetExcelDataLoad.default_load_sheet = pd.read_parquet(parquet_load_file, engine='fastparquet')

        #set the filter column read in from the parquet
        ParquetExcelDataLoad.set_filter_parquet(ParquetExcelDataLoad.parquet_filter)

    #sets the filter and loads values of the worksheets that will have its data loaded into
    def set_filter_parquet(parquet_filter):
        
        #create an empty to append the column values from the parquet filter
        worksheets = []
         
        #apply the filter to the parquet data loader file to build the list of
        #respective worksheets that will have it's data loaded into
        #there is a bug that means the filter column cannot be passed in as variable
        #it needs to be hardcoded in as an attribute of the dataframe
        ParquetExcelDataLoad.filter_column = ParquetExcelDataLoad.default_load_sheet.Question
        for fc in ParquetExcelDataLoad.filter_column:
            worksheets.append(fc)

        #pass the parameters to class variables for reuse   
        ParquetExcelDataLoad.worksheets = worksheets 

        ParquetExcelDataLoad.load_parquet_content(ParquetExcelDataLoad.excel_file, ParquetExcelDataLoad.parquet_list, ParquetExcelDataLoad.worksheets, ParquetExcelDataLoad.default_sheet_name, ParquetExcelDataLoad.custom_data_load)

    #loads the data from the parquet filter and parquet loader file into the respective
    #worksheets using the arguments from the upstream functions
    def load_parquet_content(excel_file, parquet_list, worksheets=[], default_sheet_name = 'Sheet1', custom_data_load = ''):

        #load content for the default sheet, based upon the default_sheet_name not having the default value
        #this needs to be loaded first as the default worksheet will be the active sheet once the 
        #workbook is loaded
        if default_sheet_name == 'Sheet1':
            #no data will be loaded in the default sheet
            pass
        elif default_sheet_name == excel_workbook.active.title:

            #load the excel file into memory to write the content to the worksheets
            #that have been added to the worksheets list
            excel_workbook = load_workbook(excel_file)
            
            #set up the excel writer and replace the sheet content in append mode to add the data
            excel_writer = pd.ExcelWriter(excel_file, mode="a", engine="openpyxl", if_sheet_exists="replace")
            default_sheet_parquet_load = pd.read_parquet(ParquetExcelDataLoad.parquet_load_file)
            
            #load data at default location in excel
            if custom_data_load == '':
                default_sheet_parquet_load.to_excel(excel_writer, sheet_name=default_sheet_name, index=False)
            
            #load data at custom position in excel
            elif custom_data_load != '':
                default_sheet_parquet_load.to_excel(excel_writer, sheet_name=default_sheet_name, index=False, startcol=data_load_config.cell_location_column[0], startrow=data_load_config.cell_location_row[0])

            #save the content to the Excel workbook file
            excel_writer.save()
            excel_writer.close()

        #load the excel file into memory to write the content to the worksheets
        #that have been added to the worksheets list
        excel_workbook = load_workbook(excel_file)    

        #set up the excel writer and replace the sheet content in append mode to add the data
        excel_writer = pd.ExcelWriter(excel_file, mode="a", engine="openpyxl", if_sheet_exists="replace")

        #list the parquet files and match with the worksheet names to load the content 
        #on the matching parquet leaf sub directory and worksheet name
        for pl in parquet_list:
            parent_path = pathlib.PurePath(pl)
            path_name = parent_path.parent.name
            for worksheet in worksheets:
                if path_name == worksheet:
                    parquet_content = pd.read_parquet(pl, engine='fastparquet')

                    #load data at default location in excel
                    if custom_data_load == '':
                        parquet_content.to_excel(excel_writer, sheet_name=worksheet, header=None, index=False)
                    #load data at custom position in excel
                    elif custom_data_load != '':    
                        parquet_content.to_excel(excel_writer, sheet_name=worksheet, header=None, index=False, startcol=data_load_config.cell_location_column[1], startrow=data_load_config.cell_location_row[1])

        #save the content to the Excel workbook file
        excel_writer.save()
        excel_writer.close()    