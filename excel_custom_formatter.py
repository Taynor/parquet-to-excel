import json
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Border, Side, Alignment
import excel_config

class ExcelCustomFormatter:
    def __init__(excel_file, worksheets):

        #these fields are available for the user to manipulate
        self.excel_file = excel_file

        #these fields are hidden and are for PRIVATE use only
        self.worksheets = worksheets 

    #property get and set for excel file value
    @property
    def excel_file(self):
        return self.__excel_file
    @excel_file.setter
    def excel_file(self, value):
        self.__excel_file = value    

    #property get and set for worksheets value
    @property 
    def worksheets(self):
        return self.__worksheets
    @worksheets.setter
    def worksheets(self, value):
        self.__worksheets = value

    def format_child_worksheets(excel_file):
        
        #load the workbook
        excel_workbook = load_workbook(excel_file)
        worksheet = excel_workbook.active

        #pass the parameters to class variables for reuse
        ExcelCustomFormatter.excel_file = excel_file

        #load the worksheet names to amend 
        worksheets = excel_workbook.sheetnames

        #format cell dimensions for child sheets region columns
        for ws in worksheets:
            if ws != worksheets[0]:
                worksheet = excel_workbook[ws[0:]]

                #format the width dimension of the region columns
                worksheet.column_dimensions[excel_config.child_sheet_labels[6]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[7]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[8]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[9]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[10]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[11]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[12]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[13]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[14]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[15]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[16]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[17]].width = 11.95
                worksheet.column_dimensions[excel_config.child_sheet_labels[18]].width = 11.95

                #format the height dimension of the region row
                worksheet.row_dimensions[excel_config.child_sheet_labels[21]].height = 22.50

                #format the width dimension of the total column
                worksheet.column_dimensions[excel_config.child_sheet_labels[20]].width = 20.00

                #add alignment formatting to the Total label in each child worksheet
                for row in worksheet.iter_rows(min_col=2, 
                                               min_row=10, 
                                               max_col=2, 
                                               max_row=10):
                    for cell in row:
                        if cell.value != '' or cell.value != None:
                            cell.alignment = Alignment(wrapText=True, 
                                                       vertical='center', 
                                                       horizontal='left')

                #add styling to the region column titles
                for row in worksheet.iter_rows(min_col=3, 
                                               min_row=9, 
                                               max_col=15, 
                                               max_row=9):
                    for cell in row:
                        if cell.value != '' or cell.value != None:
                            cell.alignment = Alignment(wrapText=True, 
                                                       vertical='center', 
                                                       horizontal='center')

                #cell alignment for French question on each child worksheet
                for row in worksheet.iter_rows(min_col=2, 
                                               min_row=3, 
                                               max_col=6, 
                                               max_row=3):
                    for cell in row:
                        if cell.value != '' or cell.value != None:
                            cell.alignment = Alignment(wrapText=True, 
                                                       vertical='bottom') 

                #cell alignment for English question on each child worksheet
                for row in worksheet.iter_rows(min_col=8, 
                                               min_row=3, 
                                               max_col=12, 
                                               max_row=3):
                    for cell in row:
                        if cell.value != '' or cell.value != None:
                            cell.alignment = Alignment(wrapText=True, 
                                                       vertical='bottom')  

                #custom formatting for of cell row height for row 4 of each child worksheet
                #this is default as part of the Excel template
                for row in worksheet.iter_rows(min_row=4, 
                                               max_row=4):
                    for cell in row:
                        worksheet.row_dimensions[4].height = 27.00  
                                           
        #Add cell formatting for child sheets
        excel_workbook.save(ExcelCustomFormatter.excel_file) 

        #execute the function that will format the content worksheet   
        ExcelCustomFormatter.format_content_worksheet(ExcelCustomFormatter.excel_file) 

    def format_content_worksheet(excel_file):

        #load the workbook
        excel_workbook = load_workbook(excel_file)
        worksheet = excel_workbook.active

        #pass the parameters to class variables for reuse
        ExcelCustomFormatter.excel_file = excel_file

        #load the worksheet names to amend 
        worksheets = excel_workbook.sheetnames

        #format cell width dimensions for content sheet
        for ws in worksheets:
            if ws == worksheets[0]:
                worksheet = excel_workbook[ws[0:]]
                worksheet.column_dimensions[excel_config.content_sheet_labels_dimensions[0]].width = 12.20 
                worksheet.column_dimensions[excel_config.content_sheet_labels_dimensions[1]].width = 100.00
                worksheet.column_dimensions[excel_config.content_sheet_labels_dimensions[2]].width = 100.00

                for row in worksheet.iter_rows(min_col=1, 
                                               max_col=3, 
                                               min_row=5, 
                                               max_row=5):
                    for cell in row:
                        if cell.value != '' or cell.value != None:
                            cell.alignment = Alignment(horizontal='left')

                for row in worksheet.iter_rows(min_col=2, 
                                               max_col=3, 
                                               min_row=6):
                    for cell in row:
                        if cell.value != '' or cell.value != None:
                            cell.alignment = Alignment(wrapText=True, 
                                                       vertical='bottom', 
                                                       horizontal='left')            

        #Add cell formatting for child sheets
        excel_workbook.save(excel_file)


